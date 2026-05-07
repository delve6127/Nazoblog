#!/usr/bin/env python3
"""
나조토키 리뷰 데이터 동기화 스크립트

노션 DB에서 모든 리뷰 행을 가져와 nazo_data.json의 SORT_DATA와 비교하고,
신규 리뷰 추가/점수 변경 감지/PHONETIC_MAP·COMPANY_MAP 누락 보고/검색어_맵핑.xlsx
업데이트까지 한 번에 처리한다.

사용법:
    python3 sync_nazo_data.py              # 실제 실행 (파일 저장)
    python3 sync_nazo_data.py --dry-run    # 변경사항 미리보기만 (파일 저장 X)
"""
import json
import re
import subprocess
import sys
import urllib.request
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

KST = ZoneInfo("Asia/Seoul")


def utc_iso_to_kst_date(iso_str: str) -> str:
    """노션의 UTC ISO 문자열(예: 2026-03-07T15:30:00.000Z)을
    한국시간 기준 날짜(2026-03-08)로 변환."""
    if not iso_str:
        return ""
    dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
    return dt.astimezone(KST).date().isoformat()

# ─── 설정 ───────────────────────────────────────────────────────────
DATA_SOURCE_ID = "327d25cb-2dd4-804d-97eb-000b2dd5e5c8"
TOKEN_PATH = Path.home() / ".notion_nazo_token"
PROJECT_DIR = Path(__file__).parent
JSON_PATH = PROJECT_DIR / "nazo_data.json"
XLSX_PATH = PROJECT_DIR / "검색어_맵핑.xlsx"
SITEMAP_URL = "https://nazo.monbbang.me/sitemap.xml"
NOTION_API_VERSION = "2025-09-03"

DIFF_MAP = {"아주 쉬움": 1, "쉬움": 2, "보통": 3, "어려움": 4, "아주 어려움": 5}

# 비교 대상 필드 (노션 → SORT_DATA)
COMPARE_FIELDS = ["diff", "satisfaction", "puzzle", "gimmick", "design", "language", "brand", "date"]


# ─── 유틸 ───────────────────────────────────────────────────────────
def load_token() -> str:
    if not TOKEN_PATH.exists():
        sys.exit(f"❌ 토큰 파일이 없습니다: {TOKEN_PATH}")
    return TOKEN_PATH.read_text().strip()


def notion_query_all(token: str) -> list:
    """노션 데이터베이스의 모든 행을 페이지네이션으로 가져온다."""
    url = f"https://api.notion.com/v1/data_sources/{DATA_SOURCE_ID}/query"
    headers = {
        "Authorization": f"Bearer {token}",
        "Notion-Version": NOTION_API_VERSION,
        "Content-Type": "application/json",
    }
    results = []
    next_cursor = None
    while True:
        body = {"page_size": 100}
        if next_cursor:
            body["start_cursor"] = next_cursor
        req = urllib.request.Request(
            url,
            data=json.dumps(body).encode(),
            headers=headers,
            method="POST",
        )
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
        results.extend(data.get("results", []))
        if not data.get("has_more"):
            break
        next_cursor = data.get("next_cursor")
    return results


def extract_props(page: dict) -> dict:
    """노션 페이지에서 동기화에 필요한 속성만 뽑아낸다."""
    p = page.get("properties", {})

    title_arr = p.get("이름", {}).get("title", [])
    title = "".join(t.get("plain_text", "") for t in title_arr)

    diff_text = (p.get("체감 난이도", {}).get("select") or {}).get("name")
    diff = DIFF_MAP.get(diff_text)

    brand_list = p.get("제작사", {}).get("multi_select", [])
    brand = brand_list[0]["name"] if brand_list else None

    visibility = (p.get("공개여부", {}).get("select") or {}).get("name")

    # date는 노션의 "클리어 날짜"(date 프로퍼티) — 실제 플레이 날짜.
    # 노션의 "작성일"(created_time)은 페이지를 만든 시각이라 의미가 다르다.
    clear_date = ((p.get("클리어 날짜", {}).get("date") or {}).get("start") or "")[:10]

    return {
        "title": title,
        "date": clear_date,
        "diff": diff,
        "satisfaction": p.get("개인 만족도", {}).get("number"),
        "puzzle": p.get("문제", {}).get("number"),
        "gimmick": p.get("기믹", {}).get("number"),
        "design": p.get("연출/디자인", {}).get("number"),
        "language": p.get("언어접근성", {}).get("number"),
        "brand": brand,
        "visibility": visibility,
    }


def fetch_sitemap_slugs() -> list:
    """사이트맵에서 /nazotoki-reviews/ 하위 슬러그 목록을 가져온다."""
    req = urllib.request.Request(SITEMAP_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as resp:
        text = resp.read().decode()
    locs = re.findall(r"<loc>(https?://[^<]*?/nazotoki-reviews/[^<]+)</loc>", text)
    return [loc.split("/nazotoki-reviews/", 1)[1].rstrip("/") for loc in locs]


# ─── 핵심 로직 ──────────────────────────────────────────────────────
def diff_and_apply(notion_reviews: dict, sort_data: dict, sitemap_paths: list) -> dict:
    """노션 데이터와 SORT_DATA를 비교해 변경사항을 sort_data에 반영하고,
    변경 내역을 dict로 반환한다."""
    changes = {"new": [], "updated": []}
    used_paths = {entry["url"] for entry in sort_data.values()}
    new_paths = [p for p in sitemap_paths if p not in used_paths]
    max_num = max((entry.get("num", 0) for entry in sort_data.values()), default=0)

    for title, props in notion_reviews.items():
        if title in sort_data:
            entry = sort_data[title]
            updates = []
            for field in COMPARE_FIELDS:
                new_val = props.get(field)
                if new_val is None:
                    continue
                old_val = entry.get(field)
                if old_val != new_val:
                    updates.append((field, old_val, new_val))
                    entry[field] = new_val
            if updates:
                changes["updated"].append((title, updates))
        else:
            max_num += 1
            url = new_paths.pop(0) if new_paths else "/nazotoki-reviews/UNKNOWN"
            url = url if url.startswith("/") else f"/nazotoki-reviews/{url}"
            entry = {
                "num": max_num,
                "date": props["date"],
                "diff": props["diff"],
                "satisfaction": props["satisfaction"],
                "puzzle": props["puzzle"],
                "gimmick": props["gimmick"],
                "design": props["design"],
                "language": props["language"],
                "brand": props["brand"],
                "url": url,
            }
            sort_data[title] = entry
            changes["new"].append((title, entry))
    return changes


def detect_missing(data: dict) -> tuple:
    sort_data = data["SORT_DATA"]
    phonetic_lower = {k.lower() for k in data["PHONETIC_MAP"]}
    company_lower = {k.lower() for k in data["COMPANY_MAP"]}
    missing_phonetic = [t for t in sort_data if t.lower() not in phonetic_lower]
    missing_company = sorted({
        e["brand"] for e in sort_data.values()
        if e.get("brand") and e["brand"].lower() not in company_lower
    })
    return missing_phonetic, missing_company


def _format_entry_value(entry: dict) -> str:
    """SORT_DATA 한 항목의 값(dict)을 원본 스타일에 가까운 한 줄 JSON으로."""
    body = json.dumps(entry, ensure_ascii=False, separators=(", ", ": "))
    # 중괄호 안쪽에 공백 추가: {"a":1} → { "a": 1 }
    return "{ " + body[1:-1] + " }"


def write_json_in_place(json_path: Path, sort_data: dict, updated_titles: list, new_titles: list) -> None:
    """원본 JSON 파일의 포맷을 보존한 채 SORT_DATA만 in-place 수정.

    PHONETIC_MAP과 COMPANY_MAP은 손대지 않는다. SORT_DATA의 각 항목은
    한 줄짜리 형태(`    "title": { ... }`)로 유지된다.
    """
    text = json_path.read_text(encoding="utf-8")

    # 1) 기존 엔트리 in-place 교체
    for title in updated_titles:
        title_json = json.dumps(title, ensure_ascii=False)
        new_value = _format_entry_value(sort_data[title])
        pattern = re.compile(re.escape(title_json) + r":\s*\{[^{}\n]*\}")
        text, n = pattern.subn(title_json + ": " + new_value, text, count=1)
        if n != 1:
            raise RuntimeError(f"기존 엔트리 매칭 실패: {title}")

    # 2) 새 엔트리는 SORT_DATA 닫는 } 직전에 삽입 (앞 엔트리에 콤마 추가)
    for title in new_titles:
        title_json = json.dumps(title, ensure_ascii=False)
        new_value = _format_entry_value(sort_data[title])
        new_entry_line = f"    {title_json}: {new_value}"
        pattern = re.compile(r"(\n    \"[^\"\n]+\"\s*:\s*\{[^{}\n]*\})(\n\s*\}\s*\}\s*)$")
        # re.sub 백슬래시 충돌 방지를 위해 함수형 replacement 사용
        text, n = pattern.subn(lambda m: m.group(1) + ",\n" + new_entry_line + m.group(2), text, count=1)
        if n != 1:
            raise RuntimeError(f"새 엔트리 삽입 실패: {title}")

    json_path.write_text(text, encoding="utf-8")


def read_xlsx_keywords() -> tuple:
    """검색어_맵핑.xlsx에서 키워드 칸이 채워진 항목을 읽어
    (review_keywords, brand_keywords) 두 dict를 반환.
    각 dict는 {원본 키: [키워드 리스트]} 형태."""
    try:
        import openpyxl
    except ImportError:
        return {}, {}

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    sep_seen = False
    review_keywords = {}
    brand_keywords = {}
    for row in ws.iter_rows(values_only=True):
        if row == (None, None, None):
            sep_seen = True
            continue
        if not row[1] or row[0] == "구분":
            continue
        if not row[2]:  # 키워드 칸 비어있음
            continue
        # 쉼표로 분리하고 공백 정리, 빈 항목 제거
        kws = [k.strip() for k in row[2].split(",") if k.strip()]
        if not kws:
            continue
        if not sep_seen:
            review_keywords[row[1]] = kws
        else:
            brand_keywords[row[1]] = kws
    return review_keywords, brand_keywords


def apply_xlsx_to_maps(json_path: Path, data: dict) -> dict:
    """엑셀에서 채운 키워드 중 PHONETIC_MAP/COMPANY_MAP에 아직 없는 항목을
    JSON에 in-place 추가한다. 한 줄 포맷 유지.
    반환: 작업 결과 보고용 dict"""
    review_kw, brand_kw = read_xlsx_keywords()
    sort_lower = {k.lower(): k for k in data["SORT_DATA"]}
    phonetic_lower = {k.lower() for k in data["PHONETIC_MAP"]}
    company_lower = {k.lower() for k in data["COMPANY_MAP"]}
    brand_set_lower = {
        e["brand"].lower() for e in data["SORT_DATA"].values()
        if e.get("brand")
    }

    to_add_phonetic = []  # [(json_key, [keywords]), ...]
    to_add_company = []
    skipped_no_match = []  # [(원본 엑셀키, 사유)]

    for xl_key, kws in review_kw.items():
        xl_lower = xl_key.lower()
        if xl_lower in phonetic_lower:
            continue
        if xl_lower not in sort_lower:
            skipped_no_match.append((xl_key, "PHONETIC_MAP", "SORT_DATA에 매칭 키 없음"))
            continue
        json_key = sort_lower[xl_lower]  # SORT_DATA 키 기준 (lowercase 영문 등 보존)
        to_add_phonetic.append((json_key.lower(), kws))

    for xl_key, kws in brand_kw.items():
        xl_lower = xl_key.lower()
        if xl_lower in company_lower:
            continue
        if xl_lower not in brand_set_lower:
            skipped_no_match.append((xl_key, "COMPANY_MAP", "SORT_DATA의 brand에 매칭 없음"))
            continue
        to_add_company.append((xl_lower, kws))

    if not to_add_phonetic and not to_add_company:
        return {"phonetic": [], "company": [], "skipped": skipped_no_match}

    text = json_path.read_text(encoding="utf-8")

    # PHONETIC_MAP 닫는 } 직전에 추가
    if to_add_phonetic:
        new_lines = ",\n".join(
            f'    {json.dumps(k, ensure_ascii=False)}: '
            f'{json.dumps(v, ensure_ascii=False, separators=(", ", ": "))}'
            for k, v in to_add_phonetic
        )
        # PHONETIC_MAP 마지막 엔트리 뒤에 콤마 추가 + 새 라인 삽입
        # 패턴: PHONETIC_MAP 블록의 마지막 엔트리(]) 뒤 \n  },
        pattern = re.compile(
            r'("PHONETIC_MAP"\s*:\s*\{(?:[^{}]|\{[^{}]*\})*?\])(\s*\n\s*\},)',
            re.DOTALL
        )
        replaced = [False]
        def repl(m):
            replaced[0] = True
            return m.group(1) + ",\n" + new_lines + m.group(2)
        text = pattern.sub(repl, text, count=1)
        if not replaced[0]:
            raise RuntimeError("PHONETIC_MAP 삽입 위치 매칭 실패")

    # COMPANY_MAP 닫는 } 직전에 추가
    if to_add_company:
        new_lines = ",\n".join(
            f'    {json.dumps(k, ensure_ascii=False)}: '
            f'{json.dumps(v, ensure_ascii=False, separators=(", ", ": "))}'
            for k, v in to_add_company
        )
        pattern = re.compile(
            r'("COMPANY_MAP"\s*:\s*\{(?:[^{}]|\{[^{}]*\})*?\])(\s*\n\s*\},)',
            re.DOTALL
        )
        replaced = [False]
        def repl(m):
            replaced[0] = True
            return m.group(1) + ",\n" + new_lines + m.group(2)
        text = pattern.sub(repl, text, count=1)
        if not replaced[0]:
            raise RuntimeError("COMPANY_MAP 삽입 위치 매칭 실패")

    json_path.write_text(text, encoding="utf-8")

    return {
        "phonetic": to_add_phonetic,
        "company": to_add_company,
        "skipped": skipped_no_match,
    }


def update_xlsx(new_review_titles: list, new_brand_names: list) -> None:
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl 미설치 — 엑셀 업데이트 건너뜀 (pip install openpyxl)")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    sep_row = None
    existing_review_keys = set()
    existing_brand_keys = set()
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row == (None, None, None) and sep_row is None:
            sep_row = i
            continue
        if sep_row is None and row[1]:
            existing_review_keys.add(row[1])
        elif sep_row is not None and row[1]:
            existing_brand_keys.add(row[1])

    if sep_row is None:
        print("⚠️  엑셀 분리자(빈 행)를 못 찾음 — 엑셀 업데이트 건너뜀")
        return

    for title in new_review_titles:
        if title not in existing_review_keys:
            ws.insert_rows(sep_row)
            ws.cell(row=sep_row, column=1, value="나조 이름")
            ws.cell(row=sep_row, column=2, value=title)
            sep_row += 1

    for brand in new_brand_names:
        if brand not in existing_brand_keys:
            ws.append(["브랜드", brand, None])

    wb.save(XLSX_PATH)


# ─── 출력 ───────────────────────────────────────────────────────────
def print_report(changes, missing_phonetic, missing_company, dry_run=False, map_result=None) -> None:
    print("\n" + "=" * 60)
    print("📋 동기화 결과" + (" (DRY RUN — 파일 저장 안 함)" if dry_run else ""))
    print("=" * 60)

    if changes["new"]:
        print(f"\n🆕 신규 리뷰 ({len(changes['new'])}건)")
        for title, entry in changes["new"]:
            print(f"  - {title} (#{entry['num']}, {entry['brand']}, {entry['date']})")
            print(f"      url: {entry['url']}")

    if changes["updated"]:
        print(f"\n✏️  점수 변경 ({len(changes['updated'])}건)")
        for title, fields in changes["updated"]:
            print(f"  - {title}")
            for field, old, new in fields:
                print(f"      {field}: {old} → {new}")

    if not changes["new"] and not changes["updated"]:
        print("\n✅ 변경사항 없음")

    if map_result and (map_result["phonetic"] or map_result["company"]):
        print(f"\n🔁 검색 맵에 자동 반영")
        if map_result["phonetic"]:
            print(f"  PHONETIC_MAP +{len(map_result['phonetic'])}건")
            for k, v in map_result["phonetic"]:
                print(f"    · {k}: {v}")
        if map_result["company"]:
            print(f"  COMPANY_MAP +{len(map_result['company'])}건")
            for k, v in map_result["company"]:
                print(f"    · {k}: {v}")

    if map_result and map_result["skipped"]:
        print(f"\n⚠️  엑셀 키 매칭 실패 — 수동 확인 필요 ({len(map_result['skipped'])}건)")
        for xl_key, target, reason in map_result["skipped"]:
            print(f"  · [{target}] {xl_key} — {reason}")

    if missing_phonetic:
        print(f"\n⚠️  PHONETIC_MAP 누락 ({len(missing_phonetic)}건) — 엑셀에 키워드 채우면 다음 sync에서 자동 반영")
        for t in missing_phonetic:
            print(f"  - {t}")

    if missing_company:
        print(f"\n⚠️  COMPANY_MAP 누락 ({len(missing_company)}건) — 엑셀에 키워드 채우면 다음 sync에서 자동 반영")
        for b in missing_company:
            print(f"  - {b}")

    if changes["new"]:
        print("\n📝 SEO 페이지 이름 (복붙용)")
        for title, _ in changes["new"]:
            print(f"  - {title} 리뷰 | 몬빵의 나조토키 다락방")

    print()


# ─── git 동기화 ─────────────────────────────────────────────────────
def git_pull_rebase() -> None:
    """origin/main과 먼저 동기화. 충돌하면 즉시 중단."""
    print("🔄 git pull --rebase 실행 중...")
    try:
        subprocess.run(
            ["git", "fetch", "origin", "main"],
            cwd=PROJECT_DIR, check=True, capture_output=True, text=True,
        )
        result = subprocess.run(
            ["git", "pull", "--rebase", "origin", "main"],
            cwd=PROJECT_DIR, check=True, capture_output=True, text=True,
        )
        if "Already up to date" in result.stdout or "up to date" in result.stdout.lower():
            print("   이미 최신 상태")
        else:
            print(f"   {result.stdout.strip().splitlines()[-1] if result.stdout.strip() else '동기화 완료'}")
    except subprocess.CalledProcessError as e:
        sys.exit(
            f"❌ git pull --rebase 실패 — 충돌 또는 미커밋 변경 가능성.\n"
            f"   stdout: {e.stdout}\n"
            f"   stderr: {e.stderr}\n"
            f"   수동으로 해결 후 다시 실행하세요."
        )


# ─── 메인 ───────────────────────────────────────────────────────────
def main() -> int:
    dry_run = "--dry-run" in sys.argv

    if not dry_run:
        git_pull_rebase()

    token = load_token()

    print("📡 노션 데이터베이스 조회 중...")
    pages = notion_query_all(token)

    notion_reviews = {}
    for page in pages:
        props = extract_props(page)
        if props["visibility"] != "공개":
            continue
        if not props["title"]:
            continue
        notion_reviews[props["title"]] = props
    print(f"   공개 리뷰 {len(notion_reviews)}건 수신")

    print("📡 사이트맵 조회 중...")
    sitemap_paths = [f"/nazotoki-reviews/{s}" for s in fetch_sitemap_slugs()]

    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    changes = diff_and_apply(notion_reviews, data["SORT_DATA"], sitemap_paths)
    missing_phonetic, missing_company = detect_missing(data)

    if not dry_run and (changes["new"] or changes["updated"]):
        write_json_in_place(
            JSON_PATH,
            data["SORT_DATA"],
            updated_titles=[t for t, _ in changes["updated"]],
            new_titles=[t for t, _ in changes["new"]],
        )
        update_xlsx(
            new_review_titles=[t for t, _ in changes["new"]],
            new_brand_names=missing_company,
        )

    # 엑셀에 채워진 키워드를 PHONETIC_MAP/COMPANY_MAP으로 자동 반영
    map_result = {"phonetic": [], "company": [], "skipped": []}
    if not dry_run:
        # JSON이 수정됐을 수 있으니 다시 로드
        with open(JSON_PATH, encoding="utf-8") as f:
            data = json.load(f)
        map_result = apply_xlsx_to_maps(JSON_PATH, data)
        # 미반영 항목이 사라졌는지 다시 계산
        with open(JSON_PATH, encoding="utf-8") as f:
            data = json.load(f)
        missing_phonetic, missing_company = detect_missing(data)

    print_report(changes, missing_phonetic, missing_company, dry_run=dry_run, map_result=map_result)

    if changes["new"] or changes["updated"] or map_result["phonetic"] or map_result["company"]:
        return 1  # 변경 있음 → 호출자가 git commit 처리
    return 0


if __name__ == "__main__":
    sys.exit(main())
